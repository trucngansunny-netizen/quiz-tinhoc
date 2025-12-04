# ai_tin_web.py
import os
import json
import shutil
import tempfile
import datetime
from flask import Flask, render_template, request, send_from_directory, send_file
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook

# Import module chấm tiêu chí (giữ nguyên tên file của cô)
from cham_tieuchi import (
    pretty_name_from_filename,
    load_criteria,
    grade_word,
    grade_ppt,
    grade_scratch
)

# === CẤU HÌNH ===
APP_ROOT = os.path.dirname(__file__)
STATIC_DIR = os.path.join(APP_ROOT, "static")
CRITERIA_DIR = os.path.join(APP_ROOT, "criteria")
RESULTS_DIR = os.path.join(APP_ROOT, "results")
TONGHOP_FILE = os.path.join(RESULTS_DIR, "tonghop.xlsx")   # sẽ tạo khi giáo viên tải xuống
DETAILS_FILE = os.path.join(RESULTS_DIR, "details.xlsx")   # lưu mọi nộp bài
VISIT_FILE = os.path.join(RESULTS_DIR, "visits.txt")

TEACHER_PASSWORD = "giaovien123"

ALLOWED_EXT = {"pptx", "docx", "sb3", "zip"}

CLASSES = [
    "3A1","3A2","3A3","3A4",
    "4A1","4A2","4A3","4A4","4A5",
    "5A1","5A2","5A3","5A4","5A5"
]

AVAILABLE_BY_GRADE = {
    3: ["PowerPoint"],
    4: ["Word", "PowerPoint", "Scratch"],
    5: ["Word", "Scratch"]
}

# Tạo thư mục cần thiết (nếu chưa có)
os.makedirs(STATIC_DIR, exist_ok=True)
os.makedirs(CRITERIA_DIR, exist_ok=True)
os.makedirs(RESULTS_DIR, exist_ok=True)

app = Flask(__name__, static_folder="static", template_folder="templates")
app.secret_key = os.environ.get("FLASK_SECRET", "change_this_secret_in_prod")

# === LƯỢT TRUY CẬP ===
def read_visit():
    """
    Đọc file visits.txt, trả về số nguyên. Nếu file không tồn tại -> 0.
    Lưu ý: file lưu ở RESULTS_DIR để không bị ghi đè khi deploy lại nếu folder RESULTS persist.
    """
    try:
        if not os.path.exists(VISIT_FILE):
            return 0
        with open(VISIT_FILE, "r", encoding="utf-8") as f:
            v = f.read().strip()
            return int(v or "0")
    except Exception:
        return 0

def increase_visit():
    """
    Tăng lượt truy cập +1 (an toàn, overwrite chỉ giá trị mới).
    Không reset gì khác — không tạo lại file ở startup.
    """
    try:
        current = read_visit()
        current += 1
        with open(VISIT_FILE, "w", encoding="utf-8") as f:
            f.write(str(current))
        return current
    except Exception:
        return read_visit()

# === CHỨC NĂNG LƯU DETAILS (mỗi lần 1 dòng trong sheet ChiTiet) ===
def ensure_details():
    if not os.path.exists(DETAILS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "ChiTiet"
        ws.append(["Thời gian", "Họ tên", "Lớp", "Khối", "Môn", "Tên tệp", "Tổng điểm", "Chi tiết(JSON)"])
        wb.save(DETAILS_FILE)

def save_detail_excel(student_name, class_name, grade, subject, filename, total, notes):
    ensure_details()
    wb = load_workbook(DETAILS_FILE)
    ws = wb["ChiTiet"]
    ws.append([
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        student_name,
        class_name,
        grade,
        subject,
        filename,
        float(total),
        json.dumps(notes, ensure_ascii=False)
    ])
    wb.save(DETAILS_FILE)

# === TẠO FILE TONGHOP TỪ DETAILS (khi giáo viên yêu cầu tải) ===
def build_tonghop_from_details():
    """
    Tạo file tonghop.xlsx với 1 sheet / 1 lớp từ DETAILS_FILE.
    Nếu không có row, sheet vẫn được tạo theo danh sách CLASSES.
    """
    ensure_details()
    wb_details = load_workbook(DETAILS_FILE)
    if "ChiTiet" not in wb_details.sheetnames:
        return None
    ws = wb_details["ChiTiet"]

    wb = Workbook()
    # remove default sheet if present
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # create sheets for all classes (keeps same header)
    for cls in CLASSES:
        sh = wb.create_sheet(title=cls)
        sh.append(["Thời gian", "Họ tên", "Lớp", "Khối", "Môn", "Tên tệp", "Tổng điểm", "Chi tiết(JSON)"])

    # iterate details rows and append to corresponding class sheet
    for row in list(ws.iter_rows(min_row=2, values_only=True)):
        if not row:
            continue
        time, name, class_name, grade, subject, filename, total, notes_json = row
        # sanitize class_name if None
        class_name = class_name or "Unknown"
        if class_name not in wb.sheetnames:
            ws_sheet = wb.create_sheet(title=class_name)
            ws_sheet.append(["Thời gian", "Họ tên", "Lớp", "Khối", "Môn", "Tên tệp", "Tổng điểm", "Chi tiết(JSON)"])
        ws_sheet = wb[class_name]
        ws_sheet.append([time, name, class_name, grade, subject, filename, total, notes_json])

    wb.save(TONGHOP_FILE)
    return TONGHOP_FILE

# === HỖ TRỢ KIỂM TRA FILE ===
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT

@app.route("/static/<path:filename>")
def custom_static(filename):
    return send_from_directory(STATIC_DIR, filename)

# === TRANG CHÍNH (Nộp & Chấm) ===
@app.route("/", methods=["GET", "POST"])
def index():
    # tăng lượt truy cập chỉ khi GET (mở trang)
    if request.method == "GET":
        increase_visit()

    message = None
    result = None

    if request.method == "POST":
        selected_grade = request.form.get("grade")
        selected_class = request.form.get("class")
        selected_subject = request.form.get("subject")
        uploaded = request.files.get("file")

        if not (selected_grade and selected_class and selected_subject):
            message = "Vui lòng chọn Khối, Lớp và Môn."
        elif not uploaded or uploaded.filename == "":
            message = "Vui lòng chọn tệp để nộp."
        else:
            try:
                grade_num = int(selected_grade)
            except:
                message = "Khối không hợp lệ."
                grade_num = None

            if grade_num and selected_subject not in AVAILABLE_BY_GRADE.get(grade_num, []):
                message = f"Khối {grade_num} không học môn {selected_subject}."
            else:
                filename = secure_filename(uploaded.filename)
                if not allowed_file(filename):
                    message = "Chỉ hỗ trợ .docx, .pptx, .sb3, .zip"
                else:
                    # save into a tmp dir for processing
                    tmp_dir = tempfile.mkdtemp(prefix="ai_tin_")
                    tmp_path = os.path.join(tmp_dir, filename)
                    uploaded.save(tmp_path)

                    # === THÊM MỚI: lưu bản copy thực tế của tệp vào thư mục results/<class>/<subject>/ ===
                    try:
                        dest_dir = os.path.join(RESULTS_DIR, selected_class, selected_subject)
                        os.makedirs(dest_dir, exist_ok=True)
                        # keep original filename; if exists, add timestamp suffix to avoid overwrite
                        dest_path = os.path.join(dest_dir, filename)
                        if os.path.exists(dest_path):
                            name, ext = os.path.splitext(filename)
                            stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                            dest_path = os.path.join(dest_dir, f"{name}_{stamp}{ext}")
                        shutil.copy(tmp_path, dest_path)
                    except Exception as e:
                        # không fatal — vẫn tiếp tục chấm nhưng lưu log vào message nếu cần
                        # nhưng không block chấm
                        print("Warning: không lưu được bản copy tệp học sinh:", e)

                    # load criteria
                    criteria = load_criteria(selected_subject.lower(), int(selected_grade), CRITERIA_DIR)
                    if not criteria:
                        message = f"Chưa có tiêu chí cho {selected_subject} khối {selected_grade}."
                        shutil.rmtree(tmp_dir, ignore_errors=True)
                    else:
                        notes = []
                        total = None
                        try:
                            if selected_subject == "Word":
                                total, notes = grade_word(tmp_path, criteria)
                            elif selected_subject == "PowerPoint":
                                total, notes = grade_ppt(tmp_path, criteria)
                            elif selected_subject == "Scratch":
                                total, notes = grade_scratch(tmp_path, criteria)
                        except Exception as e:
                            message = f"Lỗi khi chấm: {e}"
                            notes = [str(e)]
                            total = None

                        if total is not None:
                            student_name = pretty_name_from_filename(filename)

                            # LƯU CHI TIẾT (mỗi lần học sinh nộp) — dùng DETAILS_FILE
                            save_detail_excel(
                                student_name,
                                selected_class,
                                selected_grade,
                                selected_subject,
                                filename,
                                total,
                                notes
                            )

                            # TRẢ KẾT QUẢ CHO HỌC SINH (hiển thị chi tiết tiêu chí ✅/❌)
                            result = {
                                "student": student_name,
                                "class": selected_class,
                                "grade": selected_grade,
                                "subject": selected_subject,
                                "file": filename,
                                "total": total,
                                "details": notes
                            }
                        else:
                            message = message or "Lỗi khi chấm bài."
                        shutil.rmtree(tmp_dir, ignore_errors=True)

    avail_by_grade = {str(k): v for k, v in AVAILABLE_BY_GRADE.items()}

    return render_template(
        "index.html",
        classes=CLASSES,
        avail_by_grade=avail_by_grade,
        visit_count=read_visit(),
        result=result,
        message=message
    )

# === ROUTE BẢO MẬT CHO GIÁO VIÊN TẢI FILE TỔNG HỢP ===
@app.route("/download-tonghop", methods=["GET", "POST"])
def download_tonghop():
    if request.method == "GET":
        return """
        <html><body>
        <h3>Download file tổng hợp (giáo viên)</h3>
        <form method="post">
          Mật khẩu: <input name="pw" type="password"/>
          <button type="submit">Tải về</button>
        </form>
        <p><a href="/">Quay về trang nộp bài</a></p>
        </body></html>
        """
    pw = request.form.get("pw", "")
    if pw != TEACHER_PASSWORD:
        return "<h3>Mật khẩu sai.</h3><p><a href='/download-tonghop'>Thử lại</a></p>", 401

    # build tonghop from details (on-demand)
    if not os.path.exists(DETAILS_FILE):
        return "<h3>Chưa có dữ liệu nộp bài nào.</h3><p><a href='/'>Quay về</a></p>", 404

    path = build_tonghop_from_details()
    if not path or not os.path.exists(path):
        return "<h3>Không thể tạo file tổng hợp.</h3><p><a href='/'>Quay về</a></p>", 500

    return send_file(path, as_attachment=True, download_name=os.path.basename(path))

# === ROUTE TẢI FILE CHI TIẾT (GIÁO VIÊN) ===
@app.route("/download-details", methods=["GET", "POST"])
def download_details():
    if request.method == "GET":
        return """
        <html><body>
        <h3>Download file chi tiết (giáo viên)</h3>
        <form method="post">
          Mật khẩu: <input name="pw" type="password"/>
          <button type="submit">Tải về</button>
        </form>
        <p><a href="/">Quay về trang nộp bài</a></p>
        </body></html>
        """
    pw = request.form.get("pw", "")
    if pw != TEACHER_PASSWORD:
        return "<h3>Mật khẩu sai.</h3><p><a href='/download-details'>Thử lại</a></p>", 401

    if not os.path.exists(DETAILS_FILE):
        return "<h3>Chưa có file chi tiết nào.</h3><p><a href='/'>Quay về</a></p>", 404

    return send_file(DETAILS_FILE, as_attachment=True, download_name=os.path.basename(DETAILS_FILE))

if __name__ == "__main__":
    # app.run debug False for production; gunicorn will run it under Procfile
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
